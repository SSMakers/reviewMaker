from __future__ import annotations

from pathlib import Path
import textwrap

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[1]
ASSET_DIR = ROOT / "docs" / "assets"
SAMPLE_XLSX = ASSET_DIR / "review-writer-sample.xlsx"
EXCEL_GUIDE_PNG = ASSET_DIR / "guide-excel-template-annotated.png"
LOCAL_IMAGE_GUIDE_PNG = ASSET_DIR / "guide-local-image-upload.png"
QUICK_START_PDF = ROOT / "docs" / "Review Writer 빠른 시작 가이드.pdf"

FONT_PATH = Path("/System/Library/Fonts/AppleSDGothicNeo.ttc")
FALLBACK_FONT_PATH = Path("/System/Library/Fonts/Supplemental/AppleGothic.ttf")


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    path = FONT_PATH if FONT_PATH.exists() else FALLBACK_FONT_PATH
    index = 8 if bold and path.suffix == ".ttc" else 0
    return ImageFont.truetype(str(path), size=size, index=index)


def rounded(draw: ImageDraw.ImageDraw, box, fill, outline=None, width=1, radius=14):
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def label(draw: ImageDraw.ImageDraw, xy, text, fill="#111827", size=22, bold=False, max_width=24):
    f = font(size, bold=bold)
    lines = []
    for paragraph in text.split("\n"):
        lines.extend(textwrap.wrap(paragraph, width=max_width) or [""])
    x, y = xy
    for line in lines:
        draw.text((x, y), line, font=f, fill=fill)
        y += int(size * 1.45)
    return y


def build_sample_workbook() -> None:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "리뷰 등록 양식"

    headers = ["제목", "작성자", "리뷰내용", "별점", "날짜", "하이퍼링크", "이미지파일명"]
    rows = [
        ["", "이길남", "이미지 없이 등록할 리뷰입니다.", 4, "2026-04-10T12:43:00+09:00", "", ""],
        ["", "박성수", "이미 공개된 이미지 URL을 사용하는 리뷰입니다.", 5, "2026-04-10T12:43:00+09:00", "https://example.com/review_001.jpg", ""],
        ["", "이재성", "로컬 이미지 파일을 업로드해서 등록할 리뷰입니다.", 5, "2026-04-10T12:43:00+09:00", "test_review.jpg", ""],
        ["리뷰", "박상길", "파일명을 별도 컬럼에 적는 더 명확한 방식입니다.", 5, "2026-04-10T12:43:00+09:00", "", "test_review.jpg"],
    ]

    ws.append(headers)
    for row in rows:
        ws.append(row)

    widths = [16, 14, 48, 10, 25, 44, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif cell.row in {3, 4, 5}:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    comments = {
        "A1": "비워두면 리뷰내용 앞부분으로 제목이 자동 생성됩니다.",
        "C1": "제목과 리뷰내용이 모두 비어 있으면 해당 행은 등록되지 않습니다.",
        "D1": "1~5 사이 정수 사용을 권장합니다.",
        "F1": "URL 또는 파일명을 넣을 수 있습니다. 파일명만 넣을 때는 앱에서 이미지 폴더를 선택하세요.",
        "G1": "새 양식을 만든다면 파일명은 이 컬럼에 적는 방식이 가장 명확합니다.",
    }
    for ref, text in comments.items():
        ws[ref].comment = Comment(text, "Review Writer")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G5"
    wb.save(SAMPLE_XLSX)


def draw_excel_guide() -> None:
    img = Image.new("RGB", (1800, 1080), "#F8FAFC")
    draw = ImageDraw.Draw(img)

    label(draw, (70, 52), "엑셀은 이렇게 채우면 됩니다", size=44, bold=True, max_width=40)
    label(draw, (72, 120), "가장 중요한 규칙: 제목 또는 리뷰내용은 꼭 하나 이상 입력하고, 이미지는 URL 또는 파일명 중 하나만 적으면 됩니다.", fill="#475569", size=24, max_width=72)

    x0, y0 = 70, 210
    col_w = [170, 140, 430, 95, 235, 390, 220]
    row_h = [62, 92, 92, 92, 92]
    headers = ["제목", "작성자", "리뷰내용", "별점", "날짜", "하이퍼링크", "이미지파일명"]
    rows = [
        ["", "이길남", "이미지 없이 등록할 리뷰", "4", "2026-04-10...", "", ""],
        ["", "박성수", "URL 이미지를 사용하는 리뷰", "5", "2026-04-10...", "https://example.com/review_001.jpg", ""],
        ["", "이재성", "로컬 이미지를 업로드할 리뷰", "5", "2026-04-10...", "test_review.jpg", ""],
        ["리뷰", "박상길", "파일명을 별도 칸에 적는 방식", "5", "2026-04-10...", "", "test_review.jpg"],
    ]

    y = y0
    for r in range(len(row_h)):
        x = x0
        for c, w in enumerate(col_w):
            fill = "#111827" if r == 0 else ("#FFFFFF" if r % 2 else "#F1F5F9")
            draw.rectangle((x, y, x + w, y + row_h[r]), fill=fill, outline="#CBD5E1", width=2)
            text = headers[c] if r == 0 else rows[r - 1][c]
            color = "#FFFFFF" if r == 0 else "#111827"
            label(draw, (x + 12, y + 15), text, fill=color, size=21 if r == 0 else 19, bold=(r == 0), max_width=max(8, int(w / 13)))
            x += w
        y += row_h[r]

    # Red guide boxes.
    boxes = [
        (x0, y0, sum(col_w[:3]), row_h[0] + sum(row_h[1:])),
        (x0 + sum(col_w[:5]), y0 + row_h[0] + row_h[2], col_w[5], row_h[3]),
        (x0 + sum(col_w[:6]), y0 + row_h[0] + row_h[3], col_w[6], row_h[4]),
    ]
    for bx, by, bw, bh in boxes:
        draw.rounded_rectangle((bx - 5, by - 5, bx + bw + 5, by + bh + 5), radius=12, outline="#EF4444", width=6)

    # Callouts.
    rounded(draw, (95, 760, 530, 965), "#FFFFFF", "#EF4444", width=4)
    label(draw, (125, 790), "1. 리뷰 기본 정보", fill="#B91C1C", size=27, bold=True)
    label(draw, (125, 835), "제목은 비워도 됩니다.\n리뷰내용이 있으면 자동으로 제목을 만들어줍니다.", size=22, max_width=28)

    rounded(draw, (610, 760, 1095, 965), "#FFFFFF", "#EF4444", width=4)
    label(draw, (640, 790), "2. URL 이미지", fill="#B91C1C", size=27, bold=True)
    label(draw, (640, 835), "이미 인터넷에 올라간 이미지는\n하이퍼링크 칸에 URL을 붙여넣습니다.", size=22, max_width=31)

    rounded(draw, (1175, 760, 1705, 965), "#FFFFFF", "#EF4444", width=4)
    label(draw, (1205, 790), "3. 로컬 이미지", fill="#B91C1C", size=27, bold=True)
    label(draw, (1205, 835), "PC 폴더에 있는 이미지는 파일명만 적습니다.\n앱에서 같은 이미지 폴더를 선택하면 자동 업로드됩니다.", size=22, max_width=35)

    img.save(EXCEL_GUIDE_PNG)


def draw_local_image_guide() -> None:
    img = Image.new("RGB", (1800, 980), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    label(draw, (70, 56), "로컬 이미지를 URL 없이 등록하는 방법", size=42, bold=True, max_width=40)
    label(draw, (72, 120), "이미지를 블로그에 먼저 올릴 필요가 없습니다. 엑셀에는 파일명만 적고, 앱에서 이미지 폴더를 선택하세요.", fill="#475569", size=24, max_width=72)

    # Step cards.
    cards = [
        (80, 220, 455, 700, "1", "이미지 폴더 준비", "Test_Image 폴더 안에\n실제 이미지 파일을 넣습니다.", "test_review.jpg"),
        (520, 220, 895, 700, "2", "엑셀에 파일명 입력", "하이퍼링크 또는 이미지파일명 칸에\n파일명을 그대로 적습니다.", "test_review.jpg"),
        (960, 220, 1335, 700, "3", "앱에서 폴더 선택", "이미지 폴더 선택 버튼으로\n같은 폴더를 선택합니다.", "Test_Image"),
        (1400, 220, 1720, 700, "4", "자동 업로드", "서버가 URL을 만들고\nCafe24 리뷰 본문에 이미지를 넣습니다.", "https://.../img.jpg"),
    ]
    for x1, y1, x2, y2, num, title, body, sample in cards:
        rounded(draw, (x1, y1, x2, y2), "#F8FAFC", "#CBD5E1", width=3, radius=22)
        draw.ellipse((x1 + 28, y1 + 28, x1 + 78, y1 + 78), fill="#4F46E5")
        label(draw, (x1 + 45, y1 + 36), num, fill="#FFFFFF", size=26, bold=True)
        label(draw, (x1 + 34, y1 + 105), title, size=30, bold=True, max_width=16)
        label(draw, (x1 + 34, y1 + 165), body, fill="#334155", size=23, max_width=22)
        rounded(draw, (x1 + 34, y2 - 145, x2 - 34, y2 - 56), "#FFFFFF", "#94A3B8", width=2, radius=12)
        label(draw, (x1 + 54, y2 - 117), sample, fill="#0F172A", size=23, bold=True, max_width=20)

    # Arrows.
    for sx in [470, 910, 1350]:
        draw.line((sx, 460, sx + 34, 460), fill="#64748B", width=6)
        draw.polygon([(sx + 34, 460), (sx + 12, 446), (sx + 12, 474)], fill="#64748B")

    rounded(draw, (110, 780, 1690, 900), "#ECFDF5", "#10B981", width=3, radius=18)
    label(draw, (145, 812), "핵심: URL이 없는 이미지는 파일명만 맞으면 됩니다.", fill="#047857", size=30, bold=True, max_width=60)
    label(draw, (145, 858), "파일명과 실제 이미지 파일명이 다르면 업로드할 이미지를 찾지 못하므로, 대소문자와 확장자까지 확인하세요.", fill="#065F46", size=22, max_width=95)

    img.save(LOCAL_IMAGE_GUIDE_PNG)


def build_quick_start_pdf() -> None:
    font_file = FALLBACK_FONT_PATH if FALLBACK_FONT_PATH.exists() else FONT_PATH
    pdfmetrics.registerFont(TTFont("GuideKR", str(font_file)))

    doc = SimpleDocTemplate(
        str(QUICK_START_PDF),
        pagesize=landscape(A4),
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "TitleKR",
        parent=styles["Title"],
        fontName="GuideKR",
        fontSize=24,
        leading=30,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
    )
    body = ParagraphStyle(
        "BodyKR",
        parent=styles["BodyText"],
        fontName="GuideKR",
        fontSize=11,
        leading=17,
        textColor=colors.HexColor("#334155"),
    )
    section = ParagraphStyle(
        "SectionKR",
        parent=styles["Heading2"],
        fontName="GuideKR",
        fontSize=16,
        leading=22,
        textColor=colors.HexColor("#111827"),
        spaceBefore=4,
        spaceAfter=6,
    )

    flow = [
        Paragraph("Review Writer 빠른 시작 가이드", title),
        Paragraph("엑셀 작성부터 로컬 이미지 업로드까지, 사용자가 가장 헷갈리기 쉬운 부분만 먼저 볼 수 있도록 정리했습니다.", body),
        Spacer(1, 8),
        Paragraph("1. 엑셀 작성 예시", section),
        PdfImage(str(EXCEL_GUIDE_PNG), width=250 * mm, height=150 * mm),
        Spacer(1, 8),
        Paragraph("2. 로컬 이미지 등록 흐름", section),
        PdfImage(str(LOCAL_IMAGE_GUIDE_PNG), width=250 * mm, height=136 * mm),
        Spacer(1, 8),
        Table(
            [[
                Paragraph("<b>핵심 규칙</b>", body),
                Paragraph("이미 공개된 이미지는 URL을 넣고, PC에 있는 이미지는 파일명만 넣은 뒤 앱에서 이미지 폴더를 선택합니다.", body),
            ]],
            colWidths=[36 * mm, 210 * mm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ECFDF5")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#10B981")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]),
        ),
    ]
    doc.build(flow)


def main() -> None:
    build_sample_workbook()
    draw_excel_guide()
    draw_local_image_guide()
    build_quick_start_pdf()
    print(SAMPLE_XLSX)
    print(EXCEL_GUIDE_PNG)
    print(LOCAL_IMAGE_GUIDE_PNG)
    print(QUICK_START_PDF)


if __name__ == "__main__":
    main()
